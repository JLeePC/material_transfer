import csv
import time
import os
import openpyxl
import getpass
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def combine_workbook():
    os.chdir(r"D:\MIsys Data")

    wb = Workbook()

    dest_filename = 'Master.xlsx'

    ws = wb.active
    ws.title = "Master"

    with open('Master.csv', 'r') as csv_file:
        csv_reader = csv.reader(csv_file)

        for read in csv_reader:
            ws.append(read)

    max_row = ws.max_row

    tab = Table(displayName="Table1", ref="A1:M{}".format(max_row))

    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style

    ws.add_table(tab)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

    wb.save(filename = dest_filename)

    os.chdir(r"\\NTPV-SERVER2008\ntpv data\Justyn's MISys")

    wb.save(filename = dest_filename)
